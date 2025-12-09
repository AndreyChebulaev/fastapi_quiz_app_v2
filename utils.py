import pandas as pd
import ast
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def parse_answers(answers_str):
    """Парсит строку с ответами в список"""
    if pd.isna(answers_str):
        return [""]
    
    try:
        # Пробуем парсить как Python список
        if answers_str.startswith('"') and answers_str.endswith('"'):
            answers_str = f"[{answers_str}]"
        
        parsed = ast.literal_eval(answers_str)
        if isinstance(parsed, list):
            return [str(item).strip('"') for item in parsed]  # Убираем кавычки при парсинге
        else:
            return [str(parsed).strip('"')]
    except:
        # Если парсинг не удался, возвращаем как есть (убирая лишние кавычки)
        clean_str = str(answers_str).strip().strip('"')
        return [clean_str]

def format_answers(answers_list):
    """Форматирует список ответов в строку с кавычками"""
    if not answers_list:
        return ""
    
    # Фильтруем пустые ответы и убираем существующие кавычки
    clean_answers = []
    for answer in answers_list:
        if answer and str(answer).strip():
            clean_answer = str(answer).strip().strip('"')
            if clean_answer:  # Проверяем, не пустая ли строка после очистки
                clean_answers.append(clean_answer)
    
    if not clean_answers:
        return ""
    
    # Форматируем с кавычками
    formatted = ','.join([f'"{answer}"' for answer in clean_answers])
    return formatted

def process_excel_file(file_path):
    """Обрабатывает Excel файл и возвращает данные с сохранением формата"""
    # Читаем весь файл с помощью openpyxl для сохранения форматирования
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Получаем все данные из листа
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(list(row) if row else [])
    
    # Закрываем workbook
    wb.close()
    
    return data

def save_excel_file(file_path, data):
    """Сохраняет данные в Excel файл"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Вопросы и ответы"
    
    # Добавляем данные
    for row in data:
        ws.append(row)
    
    wb.save(file_path)
    wb.close()

def create_new_excel_file(file_path, data):
    """Создает новый Excel файл с данными БЕЗ заголовков"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Вопросы и ответы"
    
    # Добавляем данные БЕЗ заголовков - сразу вопросы и ответы
    for row in data:
        ws.append([row["question"], row["answers"]])
    
    wb.save(file_path)
    wb.close()


def check_user_permission(user_type: str, required_permission: str) -> bool:
    """
    Проверяет права доступа пользователя
    required_permission: 
    - 'admin' - доступ только для администраторов
    - 'teacher' - доступ для преподавателей и администраторов  
    - 'student' - доступ для всех пользователей
    """
    if required_permission == 'admin':
        return user_type == 'admin'
    elif required_permission == 'teacher':
        return user_type in ['teacher', 'admin']
    elif required_permission == 'student':
        return user_type in ['student', 'teacher', 'admin']
    return False