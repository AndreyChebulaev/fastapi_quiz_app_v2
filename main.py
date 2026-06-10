# fastapi_quiz_app/main.py
from fastapi import FastAPI, UploadFile, File, Form, Request, HTTPException, Depends
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
import pandas as pd
from sentence_transformers import SentenceTransformer, util
import os
import re
import torch
from typing import List, Dict, Optional, Any
import glob
import sqlite3
import hashlib
import secrets
from fastapi.middleware.cors import CORSMiddleware
import json
import time
from io import BytesIO

from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import editor
import users
from session_manager import create_session, active_sessions
from users import app as app_v2

app = FastAPI()
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")
app.mount("/v2", app_v2)
app.mount("/editor", editor.app)

# Глобальные данные
MODEL_NAME = 'all-MiniLM-L6-v2'
THRESHOLD = 0.833


def debug_log(run_id: str, hypothesis_id: str, location: str, message: str, data: Optional[Dict[str, Any]] = None) -> None:
    """Append a single NDJSON debug log line for this debug session."""
    try:
        payload = {
            "id": f"log_{int(time.time() * 1000)}",
            "timestamp": int(time.time() * 1000),
            "runId": run_id,
            "hypothesisId": hypothesis_id,
            "location": location,
            "message": message,
            "data": data or {},
        }
        with open(DEBUG_LOG_PATH, "a", encoding="utf-8") as f:
            f.write(json.dumps(payload, ensure_ascii=False) + "\n")
    except Exception:
        # Инструментация не должна ломать приложение
        pass


# endregion

# region agent log H1 – SentenceTransformer loading
try:
    model = SentenceTransformer(MODEL_NAME)
    debug_log(
        run_id="post_fix_1",
        hypothesis_id="H1",
        location="main.py:33",
        message="SentenceTransformer model loaded successfully",
        data={"model_name": MODEL_NAME},
    )
except Exception as e:
    # Подтверждённая проблема: нет доступа к huggingface.co, не роняем всё приложение
    model = None
    debug_log(
        run_id="post_fix_1",
        hypothesis_id="H1",
        location="main.py:33",
        message="SentenceTransformer model load failed (graceful fallback)",
        data={"model_name": MODEL_NAME, "error": str(e)},
    )
# endregion
questions = []
reference_answers = []
user_answers = []
all_embeddings = []
quiz_sessions = {}

# Папка для хранения загруженных файлов
UPLOAD_DIR = "uploaded_files"
os.makedirs(UPLOAD_DIR, exist_ok=True)


def validate_uploaded_excel_filename(filename: Optional[str]) -> str:
    raw_filename = (filename or "").strip()
    safe_filename = os.path.basename(raw_filename)

    if not raw_filename:
        raise HTTPException(status_code=400, detail="Не указано имя файла")
    if safe_filename != raw_filename:
        raise HTTPException(status_code=400, detail="Некорректное имя файла")
    if not safe_filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Разрешены только Excel-файлы .xlsx")

    return safe_filename


def get_uploaded_file_path(filename: str) -> str:
    safe_filename = validate_uploaded_excel_filename(filename)
    return os.path.join(UPLOAD_DIR, safe_filename)



def hash_password(password: str) -> str:
    """Хеширует пароль с использованием SHA-256 и соли"""
    salt = "quiz_system_salt"  # В реальном приложении используйте уникальную соль для каждого пользователя
    return hashlib.sha256((password + salt).encode()).hexdigest()

def verify_password(plain_password: str, hashed_password: str) -> bool:
    """Проверяет пароль"""
    return hash_password(plain_password) == hashed_password

def get_user_from_session(request: Request) -> Optional[str]:
    """Получает пользователя из сессии"""
    session_token = request.cookies.get("session_token")
    if session_token:
        from session_manager import get_user_from_session as get_session_user
        return get_session_user(session_token)
    return None

def login_required(request: Request):
    """Декоратор для проверки аутентификации"""
    user = get_user_from_session(request)
    if not user:
        raise HTTPException(status_code=401, detail="Требуется авторизация")
    return user

# Инициализация базы данных с новой структурой
def init_db():
    conn = sqlite3.connect('users.db')
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_type TEXT NOT NULL,
            last_name TEXT NOT NULL,
            first_name TEXT NOT NULL,
            middle_name TEXT,
            group_name TEXT,
            login TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS test_settings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            filename TEXT UNIQUE NOT NULL,
            time_limit_minutes INTEGER,
            available_until TEXT,
            max_attempts INTEGER,
            allowed_students TEXT,
            allowed_groups TEXT,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS test_attempts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            login TEXT NOT NULL,
            filename TEXT NOT NULL,
            attempts_count INTEGER NOT NULL DEFAULT 0,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(login, filename)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS test_results (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            login TEXT NOT NULL,
            filename TEXT NOT NULL,
            attempt_number INTEGER NOT NULL,
            total_questions INTEGER NOT NULL,
            total_correct INTEGER NOT NULL,
            percentage REAL NOT NULL,
            elapsed_seconds INTEGER,
            threshold REAL NOT NULL,
            forced_finish INTEGER NOT NULL DEFAULT 0,
            started_at TEXT,
            finished_at TEXT NOT NULL,
            user_answers_json TEXT,
            questions_json TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(login, filename, attempt_number)
        )
    ''')

    # Миграция для старых БД
    cursor.execute("PRAGMA table_info(test_settings)")
    existing_columns = {row[1] for row in cursor.fetchall()}
    if "allowed_students" not in existing_columns:
        cursor.execute("ALTER TABLE test_settings ADD COLUMN allowed_students TEXT")
    if "allowed_groups" not in existing_columns:
        cursor.execute("ALTER TABLE test_settings ADD COLUMN allowed_groups TEXT")
    if "max_attempts" not in existing_columns:
        cursor.execute("ALTER TABLE test_settings ADD COLUMN max_attempts INTEGER")

    cursor.execute("PRAGMA table_info(test_results)")
    result_columns = {row[1] for row in cursor.fetchall()}
    if "user_answers_json" not in result_columns:
        cursor.execute("ALTER TABLE test_results ADD COLUMN user_answers_json TEXT")
    if "questions_json" not in result_columns:
        cursor.execute("ALTER TABLE test_results ADD COLUMN questions_json TEXT")
    
    # Создаем тестового пользователя если его нет
    cursor.execute("SELECT COUNT(*) FROM users WHERE login = 'admin'")
    if cursor.fetchone()[0] == 0:
        cursor.execute(
            "INSERT INTO users (user_type, last_name, first_name, middle_name, group_name, login, password) VALUES (?, ?, ?, ?, ?, ?, ?)",
            ("admin", "Иванов", "Иван", "Иванович", "Администраторы", "admin", hash_password("admin123"))
        )
        print("Создан тестовый пользователь: admin / admin123")
    
    conn.commit()
    conn.close()

init_db()


def get_session_token(request: Request) -> Optional[str]:
    return request.cookies.get("session_token")


def parse_available_until(value: Optional[str]) -> Optional[datetime]:
    if not value:
        return None
    try:
        parsed_date = datetime.strptime(value, "%Y-%m-%d")
        return parsed_date + timedelta(days=1) - timedelta(seconds=1)
    except ValueError:
        return None


def format_date_ru(value: Optional[str]) -> Optional[str]:
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").strftime("%d.%m.%Y")
    except ValueError:
        return value


def format_datetime_ru(value: datetime) -> str:
    return value.strftime("%d.%m.%Y %H:%M:%S")


def format_datetime_db(value: Optional[datetime]) -> Optional[str]:
    if not value:
        return None
    return value.replace(microsecond=0).isoformat(sep=" ")


def format_duration(seconds: Optional[int]) -> str:
    if seconds is None:
        return "--"
    safe_seconds = max(0, int(seconds))
    hours = safe_seconds // 3600
    minutes = (safe_seconds % 3600) // 60
    secs = safe_seconds % 60
    if hours > 0:
        return f"{hours:02d}:{minutes:02d}:{secs:02d}"
    return f"{minutes:02d}:{secs:02d}"


def get_test_settings(filename: str) -> Dict[str, Any]:
    conn = sqlite3.connect('users.db')
    cursor = conn.cursor()
    cursor.execute(
        "SELECT time_limit_minutes, available_until, allowed_students, allowed_groups, max_attempts FROM test_settings WHERE filename = ?",
        (filename,)
    )
    row = cursor.fetchone()
    conn.close()

    time_limit_minutes = None
    available_until = None
    allowed_students = ""
    allowed_groups = ""
    max_attempts = None
    if row:
        if row[0] is not None:
            try:
                parsed_limit = int(row[0])
                time_limit_minutes = parsed_limit if parsed_limit > 0 else None
            except (TypeError, ValueError):
                time_limit_minutes = None
        if row[1]:
            available_until = str(row[1]).strip()
        if len(row) > 2 and row[2]:
            allowed_students = str(row[2]).strip()
        if len(row) > 3 and row[3]:
            allowed_groups = str(row[3]).strip()
        if len(row) > 4 and row[4] is not None:
            try:
                parsed_attempts = int(row[4])
                max_attempts = parsed_attempts if parsed_attempts > 0 else None
            except (TypeError, ValueError):
                max_attempts = None

    return {
        "time_limit_minutes": time_limit_minutes,
        "available_until": available_until,
        "allowed_students": allowed_students,
        "allowed_groups": allowed_groups,
        "max_attempts": max_attempts
    }


def parse_access_list(value: Optional[str]) -> List[str]:
    if not value:
        return []
    parts = re.split(r"[,\n;]+", str(value))
    return [p.strip() for p in parts if p.strip()]


def get_user_attempts_count(login: Optional[str], filename: str) -> int:
    if not login or not filename:
        return 0

    conn = sqlite3.connect('users.db')
    cursor = conn.cursor()
    cursor.execute(
        "SELECT attempts_count FROM test_attempts WHERE login = ? AND filename = ?",
        (login, filename)
    )
    row = cursor.fetchone()
    conn.close()

    if not row or row[0] is None:
        return 0

    try:
        return max(0, int(row[0]))
    except (TypeError, ValueError):
        return 0


def increment_user_attempts_count(login: Optional[str], filename: str) -> int:
    if not login or not filename:
        return 0

    conn = sqlite3.connect('users.db')
    cursor = conn.cursor()
    cursor.execute(
        '''
        INSERT INTO test_attempts (login, filename, attempts_count, updated_at)
        VALUES (?, ?, 1, CURRENT_TIMESTAMP)
        ON CONFLICT(login, filename) DO UPDATE SET
            attempts_count = attempts_count + 1,
            updated_at = CURRENT_TIMESTAMP
        ''',
        (login, filename)
    )
    cursor.execute(
        "SELECT attempts_count FROM test_attempts WHERE login = ? AND filename = ?",
        (login, filename)
    )
    row = cursor.fetchone()
    conn.commit()
    conn.close()

    if not row or row[0] is None:
        return 0

    try:
        return max(0, int(row[0]))
    except (TypeError, ValueError):
        return 0


def save_test_result(
    login: Optional[str],
    filename: Optional[str],
    attempt_number: int,
    total_questions: int,
    total_correct: int,
    percentage: float,
    elapsed_seconds: Optional[int],
    threshold: float,
    forced_finish: bool,
    started_at: Optional[datetime],
    finished_at: datetime,
    user_answers_list: Optional[List[str]] = None,
    questions_list: Optional[List[str]] = None
) -> None:
    if not login or not filename or attempt_number <= 0:
        return

    safe_user_answers = [str(answer) if answer is not None else "" for answer in (user_answers_list or [])]
    safe_questions = [str(question) if question is not None else "" for question in (questions_list or [])]

    conn = sqlite3.connect('users.db')
    cursor = conn.cursor()
    cursor.execute(
        '''
        INSERT INTO test_results (
            login, filename, attempt_number, total_questions, total_correct, percentage,
            elapsed_seconds, threshold, forced_finish, started_at, finished_at,
            user_answers_json, questions_json
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(login, filename, attempt_number) DO UPDATE SET
            total_questions = excluded.total_questions,
            total_correct = excluded.total_correct,
            percentage = excluded.percentage,
            elapsed_seconds = excluded.elapsed_seconds,
            threshold = excluded.threshold,
            forced_finish = excluded.forced_finish,
            started_at = excluded.started_at,
            finished_at = excluded.finished_at,
            user_answers_json = excluded.user_answers_json,
            questions_json = excluded.questions_json
        ''',
        (
            login,
            filename,
            attempt_number,
            int(total_questions),
            int(total_correct),
            float(percentage),
            int(elapsed_seconds) if elapsed_seconds is not None else None,
            float(threshold),
            1 if forced_finish else 0,
            format_datetime_db(started_at),
            format_datetime_db(finished_at),
            json.dumps(safe_user_answers, ensure_ascii=False),
            json.dumps(safe_questions, ensure_ascii=False)
        )
    )
    conn.commit()
    conn.close()


def parse_datetime_db(value: Optional[str]) -> Optional[datetime]:
    if not value:
        return None

    normalized = str(value).strip().replace("T", " ")
    if not normalized:
        return None

    try:
        return datetime.fromisoformat(normalized)
    except ValueError:
        try:
            return datetime.strptime(normalized, "%Y-%m-%d %H:%M:%S")
        except ValueError:
            return None


def parse_json_list(value: Optional[str]) -> List[str]:
    if not value:
        return []

    try:
        parsed = json.loads(value)
    except (TypeError, ValueError, json.JSONDecodeError):
        return []

    if not isinstance(parsed, list):
        return []

    result = []
    for item in parsed:
        if item is None:
            result.append("")
        elif isinstance(item, str):
            result.append(item)
        else:
            result.append(str(item))
    return result


def normalize_date_filter(value: Optional[str]) -> Optional[str]:
    if not value:
        return None
    raw = str(value).strip()
    if not raw:
        return None
    try:
        return datetime.strptime(raw, "%Y-%m-%d").strftime("%Y-%m-%d")
    except ValueError:
        return None


def normalize_search_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def tokenize_search_text(value: str) -> List[str]:
    normalized = normalize_search_text(value)
    return [token for token in re.split(r"[^\w-]+", normalized, flags=re.UNICODE) if token]


RESULT_SEARCH_STOPWORDS = {
    "покажи", "показать", "найди", "найти", "выведи", "вывести", "все", "всех",
    "результат", "результаты", "результатов", "результате", "тест", "теста", "тестов",
    "где", "у", "для", "по", "с", "со", "и", "или", "а", "но", "на", "в", "во",
    "к", "ко", "из", "от", "до", "это", "этого", "этот", "эта", "эти", "мне",
    "нужны", "нужен", "нужна", "нужно", "есть", "покажите", "выведите"
}


RUSSIAN_NUMBER_WORDS = {
    "ноль": 0,
    "один": 1, "одна": 1, "одного": 1, "одной": 1,
    "два": 2, "две": 2, "двух": 2,
    "три": 3, "трех": 3, "трёх": 3,
    "четыре": 4, "четырех": 4, "четырёх": 4,
    "пять": 5, "пяти": 5,
    "шесть": 6, "шести": 6,
    "семь": 7, "семи": 7,
    "восемь": 8, "восьми": 8,
    "девять": 9, "девяти": 9,
    "десять": 10, "десяти": 10
}


def parse_russian_number(value: str) -> Optional[float]:
    normalized = normalize_search_text(value)
    if not normalized:
        return None
    if re.fullmatch(r"\d+(?:[.,]\d+)?", normalized):
        return float(normalized.replace(",", "."))
    return float(RUSSIAN_NUMBER_WORDS[normalized]) if normalized in RUSSIAN_NUMBER_WORDS else None


def detect_comparison_operator(raw_operator: str) -> str:
    operator = normalize_search_text(raw_operator)
    if operator in {"больше", "более", "свыше", "выше", "после", ">"}:
        return "gt"
    if operator in {"не меньше", "не менее", "от", ">="}:
        return "gte"
    if operator in {"меньше", "менее", "ниже", "до", "<"}:
        return "lt"
    if operator in {"не больше", "не более", "<="}:
        return "lte"
    if operator in {"равно", "=", "ровно"}:
        return "eq"
    return "eq"


def compare_numeric_value(value: float, operator: str, threshold: float) -> bool:
    if operator == "gt":
        return value > threshold
    if operator == "gte":
        return value >= threshold
    if operator == "lt":
        return value < threshold
    if operator == "lte":
        return value <= threshold
    return value == threshold


def parse_result_search_query(search_query: str) -> Dict[str, Any]:
    normalized_query = normalize_search_text(search_query)
    cleaned_query = normalized_query
    filters: Dict[str, Dict[str, float]] = {}

    structured_patterns = [
        ("attempt_number", r"попыт\w*\s+(не\s+меньше|не\s+менее|не\s+больше|не\s+более|больше|более|свыше|выше|меньше|менее|ниже|до|от|равно|ровно|>=|<=|>|<|=)\s+([a-zа-яё0-9.,-]+)"),
        ("attempt_number", r"(не\s+меньше|не\s+менее|не\s+больше|не\s+более|больше|более|свыше|выше|меньше|менее|ниже|до|от|равно|ровно|>=|<=|>|<|=)\s+([a-zа-яё0-9.,-]+)\s+попыт\w*"),
        ("percentage_value", r"процент(?:\s+\w+){0,2}\s+(не\s+меньше|не\s+менее|не\s+больше|не\s+более|больше|более|свыше|выше|меньше|менее|ниже|до|от|равно|ровно|>=|<=|>|<|=)\s+([a-zа-яё0-9.,-]+)"),
        ("percentage_value", r"(не\s+меньше|не\s+менее|не\s+больше|не\s+более|больше|более|свыше|выше|меньше|менее|ниже|до|от|равно|ровно|>=|<=|>|<|=)\s+([a-zа-яё0-9.,-]+)\s+процент(?:\s+\w+){0,2}")
    ]

    for field_name, pattern in structured_patterns:
        match = re.search(pattern, cleaned_query, flags=re.IGNORECASE)
        if not match:
            continue
        operator = detect_comparison_operator(match.group(1))
        threshold = parse_russian_number(match.group(2))
        if threshold is None:
            continue
        filters[field_name] = {
            "operator": operator,
            "value": threshold
        }
        cleaned_query = re.sub(pattern, " ", cleaned_query, count=1, flags=re.IGNORECASE)

    semantic_tokens = [
        token for token in tokenize_search_text(cleaned_query)
        if token not in RESULT_SEARCH_STOPWORDS and not re.fullmatch(r"\d+(?:[.,]\d+)?", token)
    ]

    return {
        "original_query": normalized_query,
        "semantic_query": " ".join(semantic_tokens).strip(),
        "filters": filters
    }


def build_result_search_document(item: Dict[str, Any]) -> str:
    parts = [
        item.get("full_name", ""),
        item.get("login", ""),
        item.get("group_name", ""),
        item.get("test_name", ""),
        item.get("filename", ""),
        str(item.get("attempt_number", "")),
        str(item.get("percentage_value", "")),
        f"{item.get('total_correct', '')} {item.get('total_questions', '')}",
        item.get("finished_at_display", "")
    ]
    return normalize_search_text(" ".join(part for part in parts if part))


def compute_result_lexical_score(search_query: str, search_doc: str) -> float:
    if not search_query or not search_doc:
        return 0.0

    query_tokens = tokenize_search_text(search_query)
    doc_tokens = tokenize_search_text(search_doc)
    if not query_tokens:
        query_tokens = [normalize_search_text(search_query)]

    matched_tokens = 0
    score = 0.0

    for token in query_tokens:
        if not token:
            continue
        if token == search_doc:
            score += 1.0
            matched_tokens += 1
            continue
        if token in search_doc:
            score += 0.82
            matched_tokens += 1
            continue
        if any(doc_token.startswith(token) or token.startswith(doc_token) for doc_token in doc_tokens):
            score += 0.58
            matched_tokens += 1

    if matched_tokens == 0:
        return 0.0

    coverage_bonus = matched_tokens / max(len(query_tokens), 1)
    return score + coverage_bonus


def apply_ai_result_search(results: List[Dict[str, Any]], search_query: str) -> List[Dict[str, Any]]:
    parsed_query = parse_result_search_query(search_query)
    normalized_query = parsed_query["original_query"]
    semantic_query = parsed_query["semantic_query"]
    structured_filters = parsed_query["filters"]

    if not normalized_query:
        return results

    filtered_input_results = []
    for item in results:
        passes_filters = True
        for field_name, condition in structured_filters.items():
            raw_value = item.get(field_name)
            try:
                numeric_value = float(raw_value)
            except (TypeError, ValueError):
                passes_filters = False
                break
            if not compare_numeric_value(numeric_value, str(condition["operator"]), float(condition["value"])):
                passes_filters = False
                break
        if passes_filters:
            filtered_input_results.append(item)

    if not semantic_query:
        return filtered_input_results

    enriched_results: List[Dict[str, Any]] = []
    documents: List[str] = []

    for item in filtered_input_results:
        search_doc = build_result_search_document(item)
        lexical_score = compute_result_lexical_score(semantic_query, search_doc)

        enriched_item = dict(item)
        enriched_item["_search_doc"] = search_doc
        enriched_item["_lexical_score"] = lexical_score
        enriched_item["_semantic_score"] = 0.0
        enriched_item["_search_score"] = lexical_score
        enriched_results.append(enriched_item)
        documents.append(search_doc or " ")

    if model is not None and documents:
        try:
            query_embedding = model.encode(semantic_query, convert_to_tensor=True)
            doc_embeddings = model.encode(documents, convert_to_tensor=True)
            semantic_scores = util.cos_sim(query_embedding, doc_embeddings)[0]

            for idx, item in enumerate(enriched_results):
                semantic_score = float(semantic_scores[idx])
                item["_semantic_score"] = semantic_score
                item["_search_score"] = max(item["_lexical_score"], 0.0) * 1.35 + max(semantic_score, 0.0)
        except Exception:
            pass

    filtered_results: List[Dict[str, Any]] = []
    for item in enriched_results:
        lexical_score = float(item.get("_lexical_score", 0.0) or 0.0)
        semantic_score = float(item.get("_semantic_score", 0.0) or 0.0)
        exact_hit = semantic_query in item.get("_search_doc", "")
        token_hit = lexical_score >= 0.55
        semantic_hit = semantic_score >= 0.42 and lexical_score >= 0.2
        strong_semantic_hit = semantic_score >= 0.67

        if exact_hit or token_hit or semantic_hit or strong_semantic_hit:
            filtered_results.append(item)

    filtered_results.sort(
        key=lambda item: (
            float(item.get("_search_score", 0.0) or 0.0),
            float(item.get("_semantic_score", 0.0) or 0.0),
            float(item.get("_lexical_score", 0.0) or 0.0),
            int(item.get("id", 0) or 0)
        ),
        reverse=True
    )

    for item in filtered_results:
        item.pop("_search_doc", None)
        item.pop("_lexical_score", None)
        item.pop("_semantic_score", None)
        item.pop("_search_score", None)

    return filtered_results


def get_results_filters_from_request(request: Request) -> Dict[str, str]:
    search_query = (request.query_params.get("q") or "").strip()
    status_filter = (request.query_params.get("status") or "all").strip().lower()
    test_filter = (request.query_params.get("test") or "").strip()
    date_from = (request.query_params.get("date_from") or "").strip()
    date_to = (request.query_params.get("date_to") or "").strip()

    if status_filter not in ["all", "normal", "forced"]:
        status_filter = "all"

    return {
        "q": search_query,
        "status": status_filter,
        "test": test_filter,
        "date_from": date_from,
        "date_to": date_to
    }


def format_result_status_label(forced_finish: bool) -> str:
    return "Принудительное" if forced_finish else "Обычное"


def format_results_filter_status_label(status_filter: str) -> str:
    if status_filter == "normal":
        return "Обычное"
    if status_filter == "forced":
        return "Принудительное"
    return "Все"


def autosize_worksheet_columns(worksheet, min_width: int = 12, max_width: int = 60) -> None:
    for column_cells in worksheet.columns:
        max_length = 0
        column_index = None
        for cell in column_cells:
            column_index = cell.column
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)
        if column_index is None:
            continue
        column_letter = get_column_letter(column_index)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, min_width), max_width)


def workbook_to_excel_response(workbook: Workbook, filename: str) -> StreamingResponse:
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)

    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"'
    }
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )


def build_test_results_workbook(
    results: List[Dict[str, Any]],
    filters: Dict[str, str],
    show_user_column: bool
) -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Результаты"

    title_font = Font(bold=True, size=14)
    label_font = Font(bold=True)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="1F6AA5")

    worksheet["A1"] = "Результаты тестов"
    worksheet["A1"].font = title_font

    meta_rows = [
        ("Сформировано", format_datetime_ru(datetime.now())),
        ("Поиск", filters.get("q") or "—"),
        ("Статус", format_results_filter_status_label(filters.get("status") or "all")),
        ("Тест", filters.get("test") or "Все тесты"),
        ("Дата с", filters.get("date_from") or "—"),
        ("Дата по", filters.get("date_to") or "—"),
        ("Количество результатов", str(len(results)))
    ]

    for row_index, (label, value) in enumerate(meta_rows, start=3):
        worksheet.cell(row=row_index, column=1, value=label).font = label_font
        worksheet.cell(row=row_index, column=2, value=value)

    header_row = 11
    headers = []
    if show_user_column:
        headers.extend(["Пользователь", "Логин", "Группа"])
    headers.extend([
        "Тест",
        "Попытка",
        "Правильных ответов",
        "Всего вопросов",
        "Процент",
        "Время",
        "Завершен",
        "Статус"
    ])

    for column_index, title in enumerate(headers, start=1):
        cell = worksheet.cell(row=header_row, column=column_index, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_index, item in enumerate(results, start=header_row + 1):
        values = []
        if show_user_column:
            values.extend([
                item.get("full_name", ""),
                item.get("login", ""),
                item.get("group_name", "")
            ])
        values.extend([
            item.get("test_name", ""),
            item.get("attempt_number", ""),
            item.get("total_correct", 0),
            item.get("total_questions", 0),
            f"{item.get('percentage_display', '0.0')}%",
            item.get("elapsed_time_display", ""),
            item.get("finished_at_display", ""),
            format_result_status_label(bool(item.get("forced_finish")))
        ])

        for column_index, value in enumerate(values, start=1):
            worksheet.cell(row=row_index, column=column_index, value=value)

    worksheet.freeze_panes = f"A{header_row + 1}"
    autosize_worksheet_columns(worksheet)
    return workbook


def build_test_result_detail_workbook(result_item: Dict[str, Any], show_user_meta: bool) -> Workbook:
    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "Сводка"
    summary_sheet["A1"] = "Детали результата теста"
    summary_sheet["A1"].font = Font(bold=True, size=14)

    summary_rows = []
    if show_user_meta:
        summary_rows.extend([
            ("Пользователь", result_item.get("full_name", "")),
            ("Логин", result_item.get("login", "")),
            ("Группа", result_item.get("group_name", "") or "—")
        ])

    summary_rows.extend([
        ("Тест", result_item.get("test_name", "")),
        ("Попытка", result_item.get("attempt_number", "")),
        ("Правильных ответов", result_item.get("total_correct", 0)),
        ("Всего вопросов", result_item.get("total_questions", 0)),
        ("Процент", f"{result_item.get('percentage_display', '0.0')}%"),
        ("Статус", format_result_status_label(bool(result_item.get("forced_finish")))),
        ("Начало", result_item.get("started_at_display", "")),
        ("Завершение", result_item.get("finished_at_display", "")),
        ("Время прохождения", result_item.get("elapsed_time_display", ""))
    ])

    for row_index, (label, value) in enumerate(summary_rows, start=3):
        summary_sheet.cell(row=row_index, column=1, value=label).font = Font(bold=True)
        summary_sheet.cell(row=row_index, column=2, value=value)

    answers_start_row = len(summary_rows) + 5
    summary_sheet.cell(row=answers_start_row, column=1, value="Ответы пользователя").font = Font(bold=True, size=12)

    summary_headers_row = answers_start_row + 1
    answer_headers = ["№", "Вопрос", "Ответ пользователя", "Статус ответа"]
    for column_index, title in enumerate(answer_headers, start=1):
        cell = summary_sheet.cell(row=summary_headers_row, column=column_index, value=title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="1F6AA5")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_index, answer in enumerate(result_item.get("answer_items", []), start=summary_headers_row + 1):
        answer_text = answer.get("user_answer") or "Нет ответа"
        status_text = "Есть ответ" if answer.get("has_answer") else "Нет ответа"

        summary_sheet.cell(row=row_index, column=1, value=answer.get("index", row_index - summary_headers_row))
        summary_sheet.cell(row=row_index, column=2, value=answer.get("question", ""))
        summary_sheet.cell(row=row_index, column=3, value=answer_text)
        summary_sheet.cell(row=row_index, column=4, value=status_text)

    answers_sheet = workbook.create_sheet("Ответы")
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="1F6AA5")

    for column_index, title in enumerate(answer_headers, start=1):
        cell = answers_sheet.cell(row=1, column=column_index, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_index, answer in enumerate(result_item.get("answer_items", []), start=2):
        answer_text = answer.get("user_answer") or "Нет ответа"
        answers_sheet.cell(row=row_index, column=1, value=answer.get("index", row_index - 1))
        answers_sheet.cell(row=row_index, column=2, value=answer.get("question", ""))
        answers_sheet.cell(row=row_index, column=3, value=answer_text)
        answers_sheet.cell(
            row=row_index,
            column=4,
            value="Есть ответ" if answer.get("has_answer") else "Нет ответа"
        )

    summary_sheet.freeze_panes = "A2"
    answers_sheet.freeze_panes = "A2"
    autosize_worksheet_columns(summary_sheet)
    autosize_worksheet_columns(answers_sheet)
    return workbook


def get_test_result_tests_for_view(user_info: Optional[Dict[str, Any]]) -> List[Dict[str, str]]:
    if not user_info:
        return []

    conn = sqlite3.connect('users.db')
    cursor = conn.cursor()
    is_manager = user_info.get("user_type") in ["teacher", "admin"]

    try:
        if is_manager:
            cursor.execute(
                "SELECT DISTINCT filename FROM test_results WHERE filename IS NOT NULL AND filename != '' ORDER BY filename COLLATE NOCASE ASC"
            )
        else:
            cursor.execute(
                "SELECT DISTINCT filename FROM test_results WHERE login = ? AND filename IS NOT NULL AND filename != '' ORDER BY filename COLLATE NOCASE ASC",
                (user_info.get("login"),)
            )
        rows = cursor.fetchall()
    except sqlite3.OperationalError:
        rows = []
    finally:
        conn.close()

    options = []
    for row in rows:
        filename = str(row[0] or "").strip()
        if not filename:
            continue
        options.append({
            "filename": filename,
            "test_name": os.path.splitext(filename)[0]
        })
    return options


def get_test_results_for_view(
    user_info: Optional[Dict[str, Any]],
    limit: Optional[int] = 500,
    search_query: str = "",
    status_filter: str = "all",
    test_filter: str = "",
    date_from: str = "",
    date_to: str = ""
) -> List[Dict[str, Any]]:
    if not user_info:
        return []

    conn = sqlite3.connect('users.db')
    cursor = conn.cursor()
    is_manager = user_info.get("user_type") in ["teacher", "admin"]
    safe_search_query = normalize_search_text(search_query)
    safe_test_filter = (test_filter or "").strip()
    safe_status_filter = status_filter if status_filter in ["all", "normal", "forced"] else "all"
    safe_date_from = normalize_date_filter(date_from)
    safe_date_to = normalize_date_filter(date_to)

    try:
        base_query = '''
            SELECT
                r.id,
                r.login,
                u.last_name,
                u.first_name,
                u.middle_name,
                u.group_name,
                r.filename,
                r.attempt_number,
                r.total_correct,
                r.total_questions,
                r.percentage,
                r.elapsed_seconds,
                r.forced_finish,
                r.finished_at
            FROM test_results r
            LEFT JOIN users u ON u.login = r.login
        '''
        where_parts = []
        params: List[Any] = []

        if not is_manager:
            where_parts.append("r.login = ?")
            params.append(user_info.get("login"))

        if safe_status_filter == "normal":
            where_parts.append("r.forced_finish = 0")
        elif safe_status_filter == "forced":
            where_parts.append("r.forced_finish = 1")

        if safe_test_filter:
            where_parts.append("r.filename = ?")
            params.append(safe_test_filter)

        if safe_date_from:
            where_parts.append("datetime(r.finished_at) >= datetime(?)")
            params.append(f"{safe_date_from} 00:00:00")

        if safe_date_to:
            where_parts.append("datetime(r.finished_at) <= datetime(?)")
            params.append(f"{safe_date_to} 23:59:59")

        where_sql = ""
        if where_parts:
            where_sql = " WHERE " + " AND ".join(where_parts)

        order_part = " ORDER BY datetime(r.finished_at) DESC, r.id DESC"
        query_limit = limit if not safe_search_query else None
        if query_limit is not None:
            order_part += " LIMIT ?"
            params.append(query_limit)
        cursor.execute(base_query + where_sql + order_part, tuple(params))

        rows = cursor.fetchall()
    except sqlite3.OperationalError:
        rows = []
    finally:
        conn.close()

    results = []
    for row in rows:
        (
            result_id,
            login,
            last_name,
            first_name,
            middle_name,
            group_name,
            filename,
            attempt_number,
            total_correct,
            total_questions,
            percentage,
            elapsed_seconds,
            forced_finish,
            finished_at
        ) = row

        full_name = f"{(last_name or '').strip()} {(first_name or '').strip()}".strip()
        if middle_name:
            full_name = f"{full_name} {middle_name}".strip()
        if not full_name:
            full_name = login or "Неизвестный пользователь"

        finished_at_dt = parse_datetime_db(finished_at)
        try:
            percentage_value = float(percentage)
        except (TypeError, ValueError):
            percentage_value = 0.0

        results.append({
            "id": int(result_id or 0),
            "login": login or "",
            "full_name": full_name,
            "group_name": group_name or "",
            "filename": filename or "",
            "test_name": os.path.splitext(filename or "")[0],
            "attempt_number": int(attempt_number or 0),
            "total_correct": int(total_correct or 0),
            "total_questions": int(total_questions or 0),
            "percentage_value": percentage_value,
            "percentage_display": f"{percentage_value:.1f}",
            "elapsed_time_display": format_duration(elapsed_seconds),
            "forced_finish": bool(forced_finish),
            "finished_at_display": format_datetime_ru(finished_at_dt) if finished_at_dt else (finished_at or "—")
        })

    if safe_search_query:
        results = apply_ai_result_search(results, safe_search_query)
        if limit is not None:
            results = results[:limit]

    return results


def get_test_result_detail_for_view(user_info: Optional[Dict[str, Any]], result_id: int) -> Optional[Dict[str, Any]]:
    if not user_info or result_id <= 0:
        return None

    conn = sqlite3.connect('users.db')
    cursor = conn.cursor()
    try:
        cursor.execute(
            '''
            SELECT
                r.id,
                r.login,
                u.last_name,
                u.first_name,
                u.middle_name,
                u.group_name,
                r.filename,
                r.attempt_number,
                r.total_correct,
                r.total_questions,
                r.percentage,
                r.elapsed_seconds,
                r.forced_finish,
                r.started_at,
                r.finished_at,
                r.user_answers_json,
                r.questions_json
            FROM test_results r
            LEFT JOIN users u ON u.login = r.login
            WHERE r.id = ?
            ''',
            (result_id,)
        )
        row = cursor.fetchone()
    except sqlite3.OperationalError:
        row = None
    finally:
        conn.close()

    if not row:
        return None

    (
        db_result_id,
        login,
        last_name,
        first_name,
        middle_name,
        group_name,
        filename,
        attempt_number,
        total_correct,
        total_questions,
        percentage,
        elapsed_seconds,
        forced_finish,
        started_at,
        finished_at,
        user_answers_json,
        questions_json
    ) = row

    is_manager = user_info.get("user_type") in ["teacher", "admin"]
    current_login = user_info.get("login")
    if not is_manager and login != current_login:
        return None

    full_name = f"{(last_name or '').strip()} {(first_name or '').strip()}".strip()
    if middle_name:
        full_name = f"{full_name} {middle_name}".strip()
    if not full_name:
        full_name = login or "Неизвестный пользователь"

    questions_list = parse_json_list(questions_json)
    user_answers_list = parse_json_list(user_answers_json)

    try:
        total_questions_value = int(total_questions or 0)
    except (TypeError, ValueError):
        total_questions_value = 0

    items_count = max(total_questions_value, len(questions_list), len(user_answers_list))
    answer_items = []
    for idx in range(items_count):
        raw_question = questions_list[idx] if idx < len(questions_list) else ""
        raw_answer = user_answers_list[idx] if idx < len(user_answers_list) else ""
        question_text = raw_question.strip() if isinstance(raw_question, str) else str(raw_question or "")
        answer_text = raw_answer if isinstance(raw_answer, str) else str(raw_answer or "")

        answer_items.append({
            "index": idx + 1,
            "question": question_text or f"Вопрос {idx + 1}",
            "user_answer": answer_text.strip(),
            "has_answer": bool(answer_text.strip())
        })

    finished_at_dt = parse_datetime_db(finished_at)
    started_at_dt = parse_datetime_db(started_at)
    try:
        percentage_value = float(percentage)
    except (TypeError, ValueError):
        percentage_value = 0.0

    return {
        "id": int(db_result_id or 0),
        "login": login or "",
        "full_name": full_name,
        "group_name": group_name or "",
        "filename": filename or "",
        "test_name": os.path.splitext(filename or "")[0],
        "attempt_number": int(attempt_number or 0),
        "total_correct": int(total_correct or 0),
        "total_questions": total_questions_value,
        "percentage_display": f"{percentage_value:.1f}",
        "elapsed_time_display": format_duration(elapsed_seconds),
        "forced_finish": bool(forced_finish),
        "started_at_display": format_datetime_ru(started_at_dt) if started_at_dt else "—",
        "finished_at_display": format_datetime_ru(finished_at_dt) if finished_at_dt else (finished_at or "—"),
        "answer_items": answer_items
    }


def get_attempt_limit_status(settings: Dict[str, Any], login: Optional[str], filename: str) -> Dict[str, Any]:
    max_attempts = settings.get("max_attempts")
    attempts_used = get_user_attempts_count(login, filename) if login else 0
    limit_reached = bool(max_attempts and attempts_used >= max_attempts)
    attempts_left = None
    if max_attempts:
        attempts_left = max(0, max_attempts - attempts_used)

    return {
        "max_attempts": max_attempts,
        "attempts_used": attempts_used,
        "attempts_left": attempts_left,
        "limit_reached": limit_reached
    }


def can_user_access_test(settings: Dict[str, Any], user_info: Optional[Dict[str, Any]]) -> (bool, str):
    allowed_students = parse_access_list(settings.get("allowed_students"))
    allowed_groups = parse_access_list(settings.get("allowed_groups"))

    if not allowed_students and not allowed_groups:
        return True, ""

    if user_info and user_info.get("user_type") in ["teacher", "admin"]:
        return True, ""

    if not user_info:
        return False, "Доступ только для выбранных студентов или групп."

    login = user_info.get("login")
    group_name = user_info.get("group_name")

    if login and login in allowed_students:
        return True, ""
    if group_name and group_name in allowed_groups:
        return True, ""

    return False, "Нет доступа к тесту."


def get_quiz_restriction_status(request: Request) -> Dict[str, Any]:
    session_token = get_session_token(request)
    if not session_token:
        return {"error": None, "remaining_seconds": None, "timed_out": False, "deadline_expired": False}

    quiz_state = quiz_sessions.get(session_token)
    if not quiz_state:
        return {"error": None, "remaining_seconds": None, "timed_out": False, "deadline_expired": False}

    now = datetime.now()
    timed_out = False
    deadline_expired = False
    remaining_seconds = None

    available_until_dt = parse_available_until(quiz_state.get("available_until"))
    if available_until_dt and now > available_until_dt:
        deadline_expired = True

    time_limit_minutes = quiz_state.get("time_limit_minutes")
    started_at = quiz_state.get("started_at")
    if time_limit_minutes and started_at:
        elapsed = (now - started_at).total_seconds()
        remaining_seconds = max(0, int(time_limit_minutes * 60 - elapsed))
        if elapsed >= time_limit_minutes * 60:
            timed_out = True

    if timed_out:
        return {
            "error": "Время прохождения теста истекло.",
            "remaining_seconds": 0,
            "timed_out": True,
            "deadline_expired": deadline_expired
        }

    if deadline_expired:
        return {
            "error": "Дедлайн выполнения теста уже прошел.",
            "remaining_seconds": remaining_seconds,
            "timed_out": timed_out,
            "deadline_expired": True
        }

    return {
        "error": None,
        "remaining_seconds": remaining_seconds,
        "timed_out": False,
        "deadline_expired": False
    }


def render_select_with_error(request: Request, error_message: str):
    login = get_user_from_session(request)
    user_info = get_user_full_info(login) if login else None
    files = get_uploaded_files(user_info)
    context = get_template_context(request)
    context.update({
        "request": request,
        "files": files,
        "error": error_message
    })
    return templates.TemplateResponse("select.html", context)

def parse_quoted_strings(s):
    """Парсит строку с ответами в кавычках, разделенных запятыми"""
    return [m.group(1) for m in re.finditer(r'"([^"]*)"', s)]

def get_uploaded_files(user_info: Optional[Dict[str, Any]] = None):
    """Получает список загруженных файлов"""
    files = []
    login = user_info.get("login") if user_info else None
    for file_path in glob.glob(os.path.join(UPLOAD_DIR, "*.xlsx")):
        filename = os.path.basename(file_path)
        file_size = os.path.getsize(file_path)
        settings = get_test_settings(filename)
        has_access, _ = can_user_access_test(settings, user_info)
        if not has_access:
            continue
        attempt_limit = get_attempt_limit_status(settings, login, filename)
        available_until_dt = parse_available_until(settings.get("available_until"))
        is_available = True
        is_limit_expired = False
        unavailable_reason = ""
        if available_until_dt and datetime.now() > available_until_dt:
            is_available = False
            is_limit_expired = True
            unavailable_reason = "Дедлайн уже прошел"
        else:
            if attempt_limit["limit_reached"]:
                is_available = False
                is_limit_expired = True
                unavailable_reason = f"Лимит попыток исчерпан ({attempt_limit['attempts_used']} из {attempt_limit['max_attempts']})."
        files.append({
            "name": filename,
            "size": file_size,
            "path": file_path,
            "time_limit_minutes": settings.get("time_limit_minutes"),
            "available_until": settings.get("available_until"),
            "available_until_display": format_date_ru(settings.get("available_until")),
            "max_attempts": attempt_limit["max_attempts"],
            "attempts_used": attempt_limit["attempts_used"],
            "attempts_left": attempt_limit["attempts_left"],
            "is_available": is_available,
            "is_limit_expired": is_limit_expired,
            "unavailable_reason": unavailable_reason
        })
    return files

def get_user_by_login(login: str):
    """Получает пользователя из базы данных по логину"""
    conn = sqlite3.connect('users.db')
    cursor = conn.cursor()
    cursor.execute("""
        SELECT id, user_type, last_name, first_name, middle_name, group_name, login, password, created_at 
        FROM users WHERE login = ?
    """, (login,))
    user = cursor.fetchone()
    conn.close()
    
    if user:
        return {
            "id": user[0],
            "user_type": user[1],
            "last_name": user[2],
            "first_name": user[3],
            "middle_name": user[4],
            "group_name": user[5],
            "login": user[6],
            "password": user[7],  # Внимание: хранится в открытом виде!
            "created_at": user[8]
        }
    return None

def get_user_full_info(login: str):
    """Получает полную информацию о пользователе для отображения"""
    user = get_user_by_login(login)
    if user:
        # Формируем полное имя
        full_name = f"{user['last_name']} {user['first_name']}"
        if user['middle_name']:
            full_name += f" {user['middle_name']}"
        
        return {
            "login": user['login'],
            "full_name": full_name,
            "user_type": user['user_type'],
            "group_name": user['group_name'],
            "created_at": user['created_at']
        }
    return None


def get_user_permissions(user_type: str):
    """Возвращает доступные права пользователя"""
    permissions = {
        'can_view_tests': True,  # Все могут видеть тесты
        'can_take_tests': True,  # Все могут проходить тесты
        'can_upload_files': user_type in ['teacher', 'admin'],  # Только преподаватели и админы могут загружать файлы
        'can_delete_files': user_type in ['teacher', 'admin'],  # Только преподаватели и админы могут удалять файлы
        'can_manage_users': user_type == 'admin',  # Только админы могут управлять пользователями
        'can_edit_tests': user_type in ['teacher', 'admin'],  # Только преподаватели и админы могут редактировать тесты
    }
    return permissions

def get_template_context(request: Request):
    """Возвращает контекст для шаблонов с информацией о пользователе"""
    login = get_user_from_session(request)
    if login:
        user_info = get_user_full_info(login)
        user_permissions = get_user_permissions(user_info['user_type'])
        return {
            "user_info": user_info,
            "user_permissions": user_permissions
        }
    return {
        "user_info": None,
        "user_permissions": None
    }

@app.get("/", response_class=HTMLResponse)
def login_page(request: Request):
    """Страница входа"""
    # Если пользователь уже авторизован, перенаправляем на выбор файла
    if get_user_from_session(request):
        return RedirectResponse(url="/select", status_code=303)
    
    context = get_template_context(request)
    context["request"] = request
    return templates.TemplateResponse("login.html", context)

@app.post("/login", response_class=HTMLResponse)
def login(
    request: Request,
    username: str = Form(...),  # Изменили на username
    password: str = Form(...)
):
    """Обработка входа"""
    try:
        user = get_user_by_login(username)  # Передаем username как логин
        if not user:
            context = get_template_context(request)
            context.update({
                "request": request,
                "error": "Неверный логин или пароль"
            })
            return templates.TemplateResponse("login.html", context)
        
        # Проверяем пароль (внимание: пароль хранится в открытом виде!)
        if password != user["password"]:
            context = get_template_context(request)
            context.update({
                "request": request,
                "error": "Неверный логин или пароль"
            })
            return templates.TemplateResponse("login.html", context)
        
        # Создаем сессию и устанавливаем cookie
        from session_manager import create_session
        session_token = create_session(username)
        response = RedirectResponse(url="/select", status_code=303)
        response.set_cookie(key="session_token", value=session_token, httponly=True)
        return response
        
    except Exception as e:
        context = get_template_context(request)
        context.update({
            "request": request,
            "error": f"Ошибка входа: {str(e)}"
        })
        return templates.TemplateResponse("login.html", context)

@app.get("/select", response_class=HTMLResponse)
def select_file_page(request: Request):
    """Страница выбора файла с сервера"""
    login = get_user_from_session(request)
    if not login:
        return RedirectResponse(url="/", status_code=303)
    
    user_info = get_user_full_info(login)
    user_permissions = get_user_permissions(user_info['user_type'])
    files = get_uploaded_files(user_info)
    
    context = get_template_context(request)
    context.update({
        "request": request,
        "files": files
    })
    
    return templates.TemplateResponse("select.html", context)


@app.get("/results", response_class=HTMLResponse)
def results_page(request: Request):
    login = get_user_from_session(request)
    if not login:
        return RedirectResponse(url="/", status_code=303)

    user_info = get_user_full_info(login)
    if not user_info:
        return RedirectResponse(url="/", status_code=303)

    filters = get_results_filters_from_request(request)

    results = get_test_results_for_view(
        user_info,
        search_query=filters["q"],
        status_filter=filters["status"],
        test_filter=filters["test"],
        date_from=filters["date_from"],
        date_to=filters["date_to"]
    )
    test_options = get_test_result_tests_for_view(user_info)

    context = get_template_context(request)
    context.update({
        "request": request,
        "results": results,
        "results_count": len(results),
        "show_user_column": user_info.get("user_type") in ["teacher", "admin"],
        "test_options": test_options,
        "filters": filters
    })
    return templates.TemplateResponse("test_results.html", context)


@app.get("/results/export")
def export_results_page(request: Request):
    login = get_user_from_session(request)
    if not login:
        return RedirectResponse(url="/", status_code=303)

    user_info = get_user_full_info(login)
    if not user_info:
        return RedirectResponse(url="/", status_code=303)

    filters = get_results_filters_from_request(request)
    show_user_column = user_info.get("user_type") in ["teacher", "admin"]
    results = get_test_results_for_view(
        user_info,
        limit=None,
        search_query=filters["q"],
        status_filter=filters["status"],
        test_filter=filters["test"],
        date_from=filters["date_from"],
        date_to=filters["date_to"]
    )

    workbook = build_test_results_workbook(results, filters, show_user_column)
    export_name = f"test_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return workbook_to_excel_response(workbook, export_name)


@app.get("/results/{result_id}/export")
def export_result_detail(request: Request, result_id: int):
    login = get_user_from_session(request)
    if not login:
        return RedirectResponse(url="/", status_code=303)

    user_info = get_user_full_info(login)
    if not user_info:
        return RedirectResponse(url="/", status_code=303)

    result_item = get_test_result_detail_for_view(user_info, result_id)
    if not result_item:
        raise HTTPException(status_code=404, detail="Результат не найден")

    show_user_meta = user_info.get("user_type") in ["teacher", "admin"]
    workbook = build_test_result_detail_workbook(result_item, show_user_meta)
    safe_test_name = re.sub(r"[^A-Za-z0-9_-]+", "_", result_item.get("test_name") or "result").strip("_") or "result"
    export_name = f"test_result_{safe_test_name}_{result_item.get('attempt_number', 0)}.xlsx"
    return workbook_to_excel_response(workbook, export_name)


@app.get("/results/{result_id}", response_class=HTMLResponse)
def result_detail_page(request: Request, result_id: int):
    login = get_user_from_session(request)
    if not login:
        return RedirectResponse(url="/", status_code=303)

    user_info = get_user_full_info(login)
    if not user_info:
        return RedirectResponse(url="/", status_code=303)

    result_item = get_test_result_detail_for_view(user_info, result_id)
    if not result_item:
        raise HTTPException(status_code=404, detail="Результат не найден")

    context = get_template_context(request)
    context.update({
        "request": request,
        "result_item": result_item,
        "show_user_meta": user_info.get("user_type") in ["teacher", "admin"]
    })
    return templates.TemplateResponse("test_result_detail.html", context)


@app.post("/upload", response_class=HTMLResponse)
async def upload_file(request: Request, file: UploadFile = File(...)):
    """Загрузка нового файла на сервер"""
    # Проверяем авторизацию
    user_login = get_user_from_session(request)
    if not user_login:
        return RedirectResponse(url="/", status_code=303)
    
    # Проверяем права доступа
    user_info = get_user_full_info(user_login)
    user_permissions = get_user_permissions(user_info['user_type'])
    
    if not user_permissions['can_upload_files']:
        files = get_uploaded_files(user_info)
        context = get_template_context(request)
        context.update({
            "request": request,
            "files": files,
            "error": "У вас нет прав для загрузки файлов"
        })
        return templates.TemplateResponse("select.html", context)
    
    try:
        safe_filename = validate_uploaded_excel_filename(file.filename)

        # Сохраняем файл в папку uploaded_files
        file_path = get_uploaded_file_path(safe_filename)
        with open(file_path, "wb") as f:
            f.write(await file.read())
        
        # Возвращаем на страницу выбора файлов с сообщением об успехе
        files = get_uploaded_files(user_info)
        context = get_template_context(request)
        context.update({
            "request": request,
            "files": files,
            "message": f"Файл {safe_filename} успешно загружен!"
        })
        return templates.TemplateResponse("select.html", context)
        
    except HTTPException as exc:
        files = get_uploaded_files(user_info)
        context = get_template_context(request)
        context.update({
            "request": request,
            "files": files,
            "error": str(exc.detail)
        })
        return templates.TemplateResponse("select.html", context)

    except Exception as e:
        files = get_uploaded_files(user_info)
        context = get_template_context(request)
        context.update({
            "request": request,
            "files": files,
            "error": f"Ошибка загрузки файла: {str(e)}"
        })
        return templates.TemplateResponse("select.html", context)

async def load_quiz_data(request: Request, file_path: str):
    """Загружает данные викторины из файла и начинает тест"""
    global questions, reference_answers, user_answers, all_embeddings
    
    try:
        user_login = get_user_from_session(request)
        user_info = get_user_full_info(user_login) if user_login else None

        # region agent log H2 – загрузка теста и наличие модели
        debug_log(
            run_id="post_fix_1",
            hypothesis_id="H2",
            location="main.py:312",
            message="load_quiz_data called",
            data={"file_path": file_path, "model_is_none": model is None},
        )
        # endregion

        filename = os.path.basename(file_path)
        settings = get_test_settings(filename)
        available_until_dt = parse_available_until(settings.get("available_until"))
        if available_until_dt and datetime.now() > available_until_dt:
            return render_select_with_error(request, "Нельзя начать тест: дедлайн выполнения уже прошел.")

        has_access, access_reason = can_user_access_test(settings, user_info)
        if not has_access:
            return render_select_with_error(request, access_reason)

        attempt_limit = get_attempt_limit_status(settings, user_info.get("login") if user_info else None, filename)
        if attempt_limit["limit_reached"]:
            return render_select_with_error(
                request,
                f"Нельзя начать тест: лимит попыток исчерпан ({attempt_limit['attempts_used']} из {attempt_limit['max_attempts']})."
            )

        if model is None:
            # Модель не загружена (например, нет интернета для HuggingFace) — даём понятную ошибку пользователю
            files = get_uploaded_files(user_info)
            context = get_template_context(request)
            context.update({
                "request": request,
                "files": files,
                "error": "Модель проверки ответов не загружена (нет доступа к HuggingFace). "
                         "Обратитесь к администратору: требуется интернет или локальный кэш модели."
            })
            return templates.TemplateResponse("select.html", context)

        df = pd.read_excel(file_path, engine='openpyxl', usecols=[0,1], header=None, names=['q','a'])
        
        questions = df['q'].astype(str).tolist()
        reference_answers = [parse_quoted_strings(answers_str) for answers_str in df['a'].astype(str)]
        
        # Создаем эмбеддинги для всех эталонных ответов
        all_embeddings = []
        for answers_list in reference_answers:
            embeddings = model.encode(answers_list, convert_to_tensor=True)
            all_embeddings.append(embeddings)
        
        user_answers = []
        attempt_number = increment_user_attempts_count(user_info.get("login") if user_info else None, filename)

        session_token = get_session_token(request)
        if session_token:
            quiz_sessions[session_token] = {
                "filename": filename,
                "started_at": datetime.now(),
                "time_limit_minutes": settings.get("time_limit_minutes"),
                "available_until": settings.get("available_until"),
                "max_attempts": settings.get("max_attempts"),
                "attempt_number": attempt_number,
                "result_saved": False
            }
        
        # Перенаправляем на первый вопрос
        return RedirectResponse(url="/quiz?idx=0", status_code=303)
        
    except Exception as e:
        files = get_uploaded_files(user_info)
        context = get_template_context(request)
        context.update({
            "request": request,
            "files": files,
            "error": f"Ошибка загрузки теста: {str(e)}"
        })
        return templates.TemplateResponse("select.html", context)

@app.post("/select", response_class=HTMLResponse)
async def select_existing_file(request: Request, filename: str = Form(...)):
    """Выбор существующего файла с сервера и начало теста"""
    # Проверяем авторизацию
    user = get_user_from_session(request)
    if not user:
        return RedirectResponse(url="/", status_code=303)
    user_info = get_user_full_info(user)
    
    file_path = get_uploaded_file_path(filename)
    if not os.path.exists(file_path):
        files = get_uploaded_files(user_info)
        context = get_template_context(request)
        context.update({
            "request": request,
            "files": files,
            "error": f"Файл {filename} не найден на сервере"
        })
        return templates.TemplateResponse("select.html", context)
    
    # Загружаем данные и начинаем тест
    return await load_quiz_data(request, file_path)


@app.get("/download/{filename}")
async def download_uploaded_file(request: Request, filename: str):
    """Скачивание Excel-файла теста"""
    user_login = get_user_from_session(request)
    if not user_login:
        raise HTTPException(status_code=401, detail="Требуется авторизация")

    user_info = get_user_full_info(user_login)
    user_permissions = get_user_permissions(user_info['user_type'])

    if not user_permissions['can_edit_tests']:
        raise HTTPException(status_code=403, detail="У вас нет прав для скачивания файлов")

    safe_filename = validate_uploaded_excel_filename(filename)
    file_path = get_uploaded_file_path(safe_filename)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Файл не найден")

    return FileResponse(
        path=file_path,
        filename=safe_filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.get("/files", response_class=JSONResponse)
def get_files_list(request: Request):
    """API для получения списка файлов"""
    user = get_user_from_session(request)
    if not user:
        raise HTTPException(status_code=401, detail="Требуется авторизация")
    user_info = get_user_full_info(user)
    
    files = get_uploaded_files(user_info)
    return {"files": files}

@app.post("/delete_file")
async def delete_file(request: Request, filename: str = Form(...)):
    """Удаление файла с сервера"""
    user_login = get_user_from_session(request)
    if not user_login:
        raise HTTPException(status_code=401, detail="Требуется авторизация")
    
    # Проверяем права доступа
    user_info = get_user_full_info(user_login)
    user_permissions = get_user_permissions(user_info['user_type'])
    
    if not user_permissions['can_delete_files']:
        raise HTTPException(status_code=403, detail="У вас нет прав для удаления файлов")
    
    try:
        safe_filename = validate_uploaded_excel_filename(filename)
        file_path = get_uploaded_file_path(safe_filename)
        if os.path.exists(file_path):
            os.remove(file_path)
            return JSONResponse({"status": "success", "message": f"Файл {safe_filename} удален"})
        else:
            return JSONResponse({"status": "error", "message": "Файл не найден"})
    except Exception as e:
        return JSONResponse({"status": "error", "message": str(e)})

# Эндпоинты для управления пользователями
@app.get("/admin/users", response_class=HTMLResponse)
def admin_users_page(request: Request):
    """Страница управления пользователями"""
    login = get_user_from_session(request)
    if not login:
        return RedirectResponse(url="/", status_code=303)
    
    # Проверяем права доступа
    user_info = get_user_full_info(login)
    user_permissions = get_user_permissions(user_info['user_type'])
    
    if not user_permissions['can_manage_users']:
        # Если нет прав, перенаправляем на главную страницу тестов
        return RedirectResponse(url="/select", status_code=303)
    
    conn = sqlite3.connect('users.db')
    cursor = conn.cursor()
    cursor.execute("""
        SELECT id, user_type, last_name, first_name, middle_name, group_name, login, password, created_at 
        FROM users ORDER BY created_at DESC
    """)
    users_data = cursor.fetchall()
    conn.close()
    
    # Форматируем данные для отображения
    users = []
    for user in users_data:
        full_name = f"{user[2]} {user[3]}"
        if user[4]:
            full_name += f" {user[4]}"
        
        users.append({
            "id": user[0],
            "user_type": user[1],
            "full_name": full_name,
            "group_name": user[5],
            "login": user[6],
            "password": user[7],  # Пароль в открытом виде (небезопасно!)
            "created_at": user[8]
        })
    
    context = get_template_context(request)
    context.update({
        "request": request,
        "users": users
    })
    
    return templates.TemplateResponse("admin_users.html", context)

@app.post("/admin/add_user")
async def add_user(
    request: Request,
    user_type: str = Form(...),
    last_name: str = Form(...),
    first_name: str = Form(...),
    middle_name: str = Form(...),
    group_name: str = Form(...),
    login: str = Form(...),
    password: str = Form(...)
):
    """Добавление нового пользователя"""
    current_login = get_user_from_session(request)
    if not current_login:
        raise HTTPException(status_code=401, detail="Требуется авторизация")
    
    # Проверяем права доступа
    current_user_info = get_user_full_info(current_login)
    user_permissions = get_user_permissions(current_user_info['user_type'])
    
    if not user_permissions['can_manage_users']:
        raise HTTPException(status_code=403, detail="У вас нет прав для добавления пользователей")
    
    try:
        conn = sqlite3.connect('users.db')
        cursor = conn.cursor()
        cursor.execute(
            """INSERT INTO users 
            (user_type, last_name, first_name, middle_name, group_name, login, password) 
            VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (user_type, last_name, first_name, middle_name, group_name, login, password)
        )
        conn.commit()
        conn.close()
        return JSONResponse({"status": "success", "message": f"Пользователь {login} создан"})
    except sqlite3.IntegrityError:
        return JSONResponse({"status": "error", "message": "Пользователь с таким логином уже существует"})
    except Exception as e:
        return JSONResponse({"status": "error", "message": str(e)})

@app.post("/admin/delete_user")
async def delete_user(request: Request, user_id: int = Form(...)):
    """Удаление пользователя"""
    current_login = get_user_from_session(request)
    if not current_login:
        raise HTTPException(status_code=401, detail="Требуется авторизация")
    
    # Проверяем права доступа
    current_user_info = get_user_full_info(current_login)
    user_permissions = get_user_permissions(current_user_info['user_type'])
    
    if not user_permissions['can_manage_users']:
        raise HTTPException(status_code=403, detail="У вас нет прав для удаления пользователей")
    
    try:
        conn = sqlite3.connect('users.db')
        cursor = conn.cursor()
        cursor.execute("DELETE FROM users WHERE id = ?", (user_id,))
        conn.commit()
        conn.close()
        return JSONResponse({"status": "success", "message": "Пользователь удален"})
    except Exception as e:
        return JSONResponse({"status": "error", "message": str(e)})

# Эндпоинты теста (требуют авторизации)
@app.get("/quiz", response_class=HTMLResponse)
def quiz_form(request: Request, idx: int = 0):
    login = get_user_from_session(request)
    if not login:
        return RedirectResponse(url="/", status_code=303)
    
    user_info = get_user_full_info(login)
    user_permissions = get_user_permissions(user_info['user_type'])

    restriction_status = get_quiz_restriction_status(request)
    if restriction_status["error"]:
        return render_select_with_error(request, restriction_status["error"])
    
    if idx >= len(questions):
        return RedirectResponse(url="/final_results", status_code=303)
    
    current_answer = user_answers[idx] if idx < len(user_answers) else ""
    
    context = get_template_context(request)
    context.update({
        "request": request,
        "question": questions[idx],
        "idx": idx,
        "current_answer": current_answer,
        "total_questions": len(questions),
        "questions": questions,
        "remaining_seconds": restriction_status["remaining_seconds"]
    })
    
    return templates.TemplateResponse("quiz.html", context)

@app.post("/answer")
async def save_answer(request: Request, idx: int = Form(...), user_answer: str = Form(...)):
    user = get_user_from_session(request)
    if not user:
        raise HTTPException(status_code=401, detail="Требуется авторизация")
    
    restriction_status = get_quiz_restriction_status(request)
    if restriction_status["error"]:
        raise HTTPException(status_code=403, detail=restriction_status["error"])

    global user_answers
    try:
        while len(user_answers) <= idx:
            user_answers.append("")
            
        user_answers[idx] = user_answer.strip()
        return JSONResponse({"status": "success"})
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/navigate")
async def navigate_question(request: Request, current_idx: int = Form(...), direction: str = Form(...)):
    user = get_user_from_session(request)
    if not user:
        raise HTTPException(status_code=401, detail="Требуется авторизация")
    
    restriction_status = get_quiz_restriction_status(request)
    if restriction_status["error"]:
        raise HTTPException(status_code=403, detail=restriction_status["error"])

    try:
        if direction == "next":
            new_idx = current_idx + 1
        else:
            new_idx = current_idx - 1
        
        # Проверяем границы
        if new_idx < 0:
            new_idx = 0
        elif new_idx >= len(questions):
            # Если пытаемся перейти за последний вопрос - перенаправляем на завершение
            return RedirectResponse(url="/final_results", status_code=303)
        
        return RedirectResponse(url=f"/quiz?idx={new_idx}", status_code=303)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

async def check_test_completion(request: Request):
    """Общая логика проверки завершения теста"""
    user = get_user_from_session(request)
    if not user:
        raise HTTPException(status_code=401, detail="Требуется авторизация")
    
    restriction_status = get_quiz_restriction_status(request)
    if restriction_status["error"]:
        raise HTTPException(status_code=403, detail=restriction_status["error"])

    global user_answers, questions
    unanswered = []
    
    for i in range(len(questions)):
        if i >= len(user_answers) or not user_answers[i].strip():
            unanswered.append(i + 1)
    
    return {
        "completed": len(unanswered) == 0,
        "unanswered": unanswered,
        "total_questions": len(questions),
        "answered_count": sum(1 for ans in user_answers if ans and ans.strip())
    }

@app.post("/check_completion")
async def check_test_completion_post(request: Request): 
    """POST версия для проверки завершения"""
    return await check_test_completion(request)

@app.get("/check_completion")
async def check_test_completion_get(request: Request):
    """GET версия для проверки завершения"""
    return await check_test_completion(request)

@app.get("/final_results", response_class=HTMLResponse)
def show_final_results(request: Request):
    user = get_user_from_session(request)
    if not user:
        return RedirectResponse(url="/", status_code=303)
    
    user_info = get_user_full_info(user)
    user_permissions = get_user_permissions(user_info['user_type'])
    global user_answers, questions

    session_token = get_session_token(request)
    quiz_state = quiz_sessions.get(session_token) if session_token else None
    now = datetime.now()
    started_at = quiz_state.get("started_at") if quiz_state else None
    finished_at = now
    if quiz_state:
        if quiz_state.get("finished_at"):
            finished_at = quiz_state["finished_at"]
        else:
            quiz_state["finished_at"] = now
            finished_at = now

    elapsed_seconds = None
    if started_at:
        elapsed_seconds = max(0, int((finished_at - started_at).total_seconds()))
    restriction_status = get_quiz_restriction_status(request)
    force_finish = restriction_status["timed_out"] or restriction_status["deadline_expired"]
    
    if len(user_answers) < len(questions) and not force_finish:
        for i in range(len(questions)):
            if i >= len(user_answers) or not user_answers[i].strip():
                context = get_template_context(request)
                context.update({
                    "request": request,
                    "unanswered_index": i,
                    "total_questions": len(questions),
                    "answered_count": sum(1 for ans in user_answers if ans and ans.strip())
                })
                return templates.TemplateResponse("complete_all.html", context)
    
    answers_for_scoring = list(user_answers)
    if len(answers_for_scoring) < len(questions):
        answers_for_scoring.extend([""] * (len(questions) - len(answers_for_scoring)))

    user_embeddings = model.encode(answers_for_scoring, convert_to_tensor=True)
    results = []
    total_correct = 0
    
    for i, (user_emb, user_answer) in enumerate(zip(user_embeddings, answers_for_scoring)):
        if not user_answer.strip():
            results.append({
                "question": questions[i],
                "user_answer": "",
                "is_correct": False,
                "score": "0.00",
                "best_reference_answer": reference_answers[i][0] if reference_answers[i] else "",
                "reference_answers": reference_answers[i],
                "max_similarity": 0.0
            })
            continue

        question_embeddings = all_embeddings[i]
        similarities = util.cos_sim(user_emb, question_embeddings)[0]
        
        max_sim_idx = int(similarities.argmax())
        max_similarity = float(similarities[max_sim_idx])
        best_reference_answer = reference_answers[i][max_sim_idx]
        
        is_correct = max_similarity >= THRESHOLD
        if is_correct:
            total_correct += 1
        
        results.append({
            "question": questions[i],
            "user_answer": user_answer,
            "is_correct": is_correct,
            "score": f"{max_similarity:.2f}",
            "best_reference_answer": best_reference_answer,
            "reference_answers": reference_answers[i],
            "max_similarity": max_similarity
        })
    
    total_questions = len(questions)
    percentage = (total_correct / total_questions) * 100 if total_questions > 0 else 0

    if quiz_state and not quiz_state.get("result_saved"):
        save_test_result(
            login=user,
            filename=quiz_state.get("filename"),
            attempt_number=int(quiz_state.get("attempt_number") or 0),
            total_questions=total_questions,
            total_correct=total_correct,
            percentage=percentage,
            elapsed_seconds=elapsed_seconds,
            threshold=THRESHOLD,
            forced_finish=force_finish,
            started_at=started_at,
            finished_at=finished_at,
            user_answers_list=answers_for_scoring,
            questions_list=questions
        )
        quiz_state["result_saved"] = True

    context = get_template_context(request)
    context.update({
        "request": request,
        "results": results,
        "total_correct": total_correct,
        "total_questions": total_questions,
        "percentage": f"{percentage:.1f}",
        "threshold": THRESHOLD,
        "restriction_message": restriction_status["error"] if force_finish else None,
        "completed_at_display": format_datetime_ru(finished_at),
        "elapsed_time_display": format_duration(elapsed_seconds)
    })
    
    return templates.TemplateResponse("final_results.html", context)

@app.get("/logout")
def logout():
    """Выход из системы"""
    response = RedirectResponse(url="/", status_code=303)
    response.delete_cookie("session_token")
    return response

@app.get("/app")
def redirect_to_editor(request: Request):
    """Перенаправление на редактор тестов"""
    user = get_user_from_session(request)
    if not user:
        return RedirectResponse(url="/", status_code=303)
    
    user_info = get_user_full_info(user)
    user_permissions = get_user_permissions(user_info['user_type'])
    
    if not user_permissions['can_edit_tests']:
        return RedirectResponse(url="/select", status_code=303)
    
    return RedirectResponse(url="/editor", status_code=307)
# Для запуска:
# uvicorn main:app --reload

